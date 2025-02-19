import logging
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import re


class ExcelSorting:
    def __init__(self):
        pass

    @staticmethod
    def create_table(worksheet, table_name, matches, row_index, start_col_index):
        sanitized_name = re.sub(r"[^a-zA-Z0-9_]", "_", table_name)
        # row_index is the first data row
        start_row = row_index - 1  # Header row
        end_row = start_row + len(matches)  # Includes header and data rows

        table_range = f"{get_column_letter(start_col_index)}{start_row}:{get_column_letter(start_col_index + 5)}{end_row}"
        table = Table(displayName=sanitized_name, ref=table_range)
        table_style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = table_style
        worksheet.add_table(table)
