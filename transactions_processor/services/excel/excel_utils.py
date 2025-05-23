import logging
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from transactions_processor.models.transaction import Transaction


class ExcelHelpers:

    @staticmethod
    def _write_transaction(
        worksheet: Worksheet,
        row_index,
        column_start_index,
        transaction: Transaction,
        format=None,
    ):
        logger = logging.getLogger(__name__)
        logger.debug(
            f" writing transaction {transaction} in excel's row {row_index} and from column index {column_start_index}"
        )
        worksheet[f"{get_column_letter(column_start_index)}{row_index}"].value = transaction.uuid
        worksheet[f"{get_column_letter(column_start_index+1)}{row_index}"].value = transaction.date.strftime("%m-%d-%Y")
        worksheet[f"{get_column_letter(column_start_index+2)}{row_index}"].value = transaction.description
        worksheet[f"{get_column_letter(column_start_index+3)}{row_index}"].value = transaction.amount
        if format is not None:
            worksheet[f"{get_column_letter(column_start_index)}{row_index}"].fill = format
            worksheet[f"{get_column_letter(column_start_index+1)}{row_index}"].fill = format
            worksheet[f"{get_column_letter(column_start_index+2)}{row_index}"].fill = format
            worksheet[f"{get_column_letter(column_start_index+3)}{row_index}"].fill = format
        worksheet[f"E{row_index}"].fill = PatternFill("solid", fgColor="000000")

    @staticmethod
    def _setup_excel(
         worksheet: Worksheet, bank_name="Bank", system_name="PharmBills System"
    ):
        ExcelHelpers._setup_headers(
            worksheet=worksheet,
            bank_name=bank_name,
            system_name=system_name,
        )
        green_fill = PatternFill(fgColor="C6EFCE", fill_type="solid")
        color_formats = {
            "pink": "FFD1DC",
            "fuchsia": "FF77FF",
            "purple": "A291FB",
            "medium_orchid": "D1B3EB",
            "dark_violet": "A64DFF",
            "indigo": "6E6EF9",
            "navy": "3F4DFA",
            "blue": "3F6DFA",
            "dark_blue": "4169E1",
            "deep_sky_blue": "00BFFF",
            "turquoise": "60E0D0",
            "aqua": "5FFBFF",
            "teal": "20B2AA",
            "dark_green": "66C966",
            "olive": "AABD75",
            "light_green": "ACF0AC",
            "lime": "66FF66",
            "yellow": "FFFF66",
            "gold": "FFD966",
            "orange": "FFB347",
        }
        color_fills = {
            index: PatternFill("solid",fgColor=code, fill_type="solid")
            for index, (name, code) in enumerate(color_formats.items())
        }
        return green_fill, color_fills
    
    @staticmethod
    def _setup_headers(worksheet: Worksheet, bank_name="Bank", system_name="PharmBills System"):
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        merged_header_fill = PatternFill("solid", fgColor="DDEBF7")

        # Setting column widths
        columns = ["A", "B", "C", "D", "F", "G", "H", "I"]
        for col in columns:
            worksheet.column_dimensions[col].width = 20
        worksheet.column_dimensions["E"].width = 3  # Empty column for separation

        # Merging for main header titles and setting their values
        worksheet.merge_cells("A1:D1")
        worksheet["A1"].value = system_name
        worksheet.merge_cells("F1:I1")
        worksheet["F1"].value = bank_name

        # Applying styles to merged headers
        for cell in ["A1", "F1"]:
            worksheet[cell].fill = merged_header_fill
            worksheet[cell].font = header_font
            worksheet[cell].alignment = header_alignment

        # Setting individual headers below main titles
        headers = ["UID", "Date", "Description", "Amount"]
        for i, header in enumerate(headers, start=1):
            cell = f"{get_column_letter(i)}2"
            worksheet[cell].value = header
            worksheet[cell].font = Font(bold=True)
            worksheet[cell].alignment = Alignment(horizontal="center", vertical="center")

            # Repeat for bank headers, offset by 5 columns due to separation
            bank_cell = f"{get_column_letter(i + 5)}2"
            worksheet[bank_cell].value = header
            worksheet[bank_cell].font = Font(bold=True)
            worksheet[bank_cell].alignment = Alignment(horizontal="center", vertical="center")

        # Freeze panes
        # worksheet[f"E1"].fill = PatternFill("solid", fgColor="000000")
        # worksheet[f"E2"].fill = PatternFill("solid", fgColor="000000")
        worksheet.freeze_panes = "A3"
    
    @staticmethod
    def _create_title_row(worksheet, title, column_start_index, row_index):
        start_cell = f"{get_column_letter(column_start_index)}{row_index}"
        stop_cell = f"{get_column_letter(column_start_index + 3)}{row_index}"
        worksheet.merge_cells(f"{start_cell}:{stop_cell}")
        worksheet[start_cell].value = title
        worksheet[start_cell].font = Font(bold=True, color="FFFFFF")
        worksheet[start_cell].alignment = Alignment(horizontal="center", vertical="center")
        worksheet[start_cell].fill = PatternFill("solid", fgColor="000000")
        worksheet[f"E{row_index}"].fill = PatternFill("solid", fgColor="000000")
        row_index += 1

        headers = ["UID", "Date", "Description", "Amount"]
        for i, header in enumerate(headers, start=1):
            cell = f"{get_column_letter(i+column_start_index-1)}{row_index}"
            worksheet[cell].value = header
            worksheet[cell].font = Font(bold=True)
            worksheet[cell].alignment = Alignment(horizontal="center", vertical="center")

        row_index += 1
        return row_index

    
